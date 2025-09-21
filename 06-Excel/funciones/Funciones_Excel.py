import openpyxl

class Funexcel:
    def __init__(self, driver=None):
        self.driver = driver

    def getRowCount(self, file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_row

    def getColumnCount(self, file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_column

    def readData(self, file, sheetName, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, file, sheetName, rownum, colnum, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        workbook.save(file)
